import json
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from PIL import Image as PILImage, ImageEnhance, ImageFilter
import pytesseract

# Path to the JSON file
json_path = "resources.json"
pytesseract.pytesseract.tesseract_cmd = 'C:/Program Files/Tesseract-OCR/tesseract.exe'  # your path may be different

# Function to extract the quantity from an image using OCR
def extract_quantity(image_path):
    img = PILImage.open(image_path)
    
    # Convert image to grayscale
    img = img.convert('L')
    
    # Crop the image to focus on the quantity region (manual adjustment might be needed)
    # Adjust the crop box values as per your image structure
    width, height = img.size
    crop_box = (0, 0, width, height // 4)  # Example values, adjust accordingly
    img = img.crop(crop_box)
    
    # Enhance the image contrast
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(2)
    
    # Apply additional image filtering
    img = img.filter(ImageFilter.SHARPEN)
    
    # Use OCR to extract text
    text = pytesseract.image_to_string(img, config='--psm 6 digits')
    print(f"OCR text for {image_path}: {text}")
    try:
        # Extract the number from the text
        quantity = int(''.join(filter(str.isdigit, text)))
    except (IndexError, ValueError):
        quantity = 0
    return quantity

# Read resources from JSON file
with open(json_path, 'r') as f:
    data = json.load(f)

resources = data['resources']

# Create a new Excel workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Resources"

# Set column widths
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 15

# Set row height

# Add headers
ws.append(["Resource", "Image", "Quantity"])

# Add resources to the Excel file
for i, resource in enumerate(resources, start=2):
    name = resource['name']
    img_path = resource['image']
    qty = extract_quantity(img_path)

    ws.cell(row=i, column=1, value=name)
    img = Image(img_path)
    img.height = 40  # Adjust the image size if needed
    img.width = 40
    ws.add_image(img, f'B{i}')
    ws.cell(row=i, column=3, value=qty)
    ws.cell(row=i, column=3).alignment = Alignment(horizontal='center')

# Save the workbook
excel_path = "resource_management.xlsx"
wb.save(excel_path)